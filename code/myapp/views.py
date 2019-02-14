from django.shortcuts import render
import openpyxl
import requests, json 

#file download package
from django.http import HttpResponse
from wsgiref.util import FileWrapper



# enter your api key here 
api_key = 'AIzaSyDBe7gTDaWPzdGkNRPGmnlBBm_AfszqNIo'
  
# url variable store url 
url = 'https://maps.googleapis.com/maps/api/geocode/json?'
  

def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)

        #write code 
        # to create a new blank Workbook object 
        kb = openpyxl.Workbook() 
        # Get workbook active sheet   
        # from the active attribute.  
        sheet = kb.active 

        m = 1
        excel_data_lat_long = list()
        # reading a cell
        #print(worksheet["A1"].value)
        for cell in worksheet['D']:
            if cell.value == 'Address':
                pass
            else:
                #print(cell.value)
                place = cell.value
                res_ob = requests.get(url+'&address='+place+'&key='+api_key) 
                x = res_ob.json() 
                #seperate data lat and lng 
                lat = x['results'][0]['geometry']['location']['lat']
                lng = x['results'][0]['geometry']['location']['lng']
                # print the vale of x 
                excel_data_lat_long.append(place+' = '+str(lat)+','+str(lng))
                #sheet A column add value using passing  integer value
                sheet['A'+str(m)].value = place
                sheet['B'+str(m)].value = str(lat)+','+str(lng)
                m = m + 1
                print("--------------")
                print(lat)
            kb.save('demo.xlsx')


        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                #print(cell.value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data,"excel_data_lat_long":excel_data_lat_long,'file_name':'demo.xlsx'})




def download_pdf(request):
    filename = "demo.xlsx" # this is the file people must download
    with open(filename, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + filename
        response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
        return response